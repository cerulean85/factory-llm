# -*- coding: utf-8 -*-
"""
update_fs_xlsx.py
Updates Functional_Specification_v1.0.xlsx based on URS v0.1 (urs_v0_1.docx)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_functional_specification():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Common Styles
    font_title = Font(name="Malgun Gothic", size=14, bold=True, color="1F497D")
    font_section = Font(name="Malgun Gothic", size=11, bold=True, color="1F497D")
    font_header = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Malgun Gothic", size=9.5, color="000000")
    font_data_bold = Font(name="Malgun Gothic", size=9.5, bold=True, color="000000")
    font_code = Font(name="Consolas", size=9.0, color="000000")
    font_desc = Font(name="Malgun Gothic", size=9.0, color="555555")

    fill_navy = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_sub_header = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_accent = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")

    thin_gray = Side(style="thin", color="D9D9D9")
    thick_navy = Side(style="medium", color="1F497D")
    double_navy = Side(style="double", color="1F497D")

    border_cell = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    border_header = Border(left=thin_gray, right=thin_gray, top=thick_navy, bottom=thick_navy)
    border_table_bottom = Border(bottom=thick_navy)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # =========================================================================
    # SHEET 1: 1. 문서정보 및 개정이력
    # =========================================================================
    ws1 = wb.create_sheet(title="1. 문서정보 및 개정이력")
    ws1.views.sheetView[0].showGridLines = True

    ws1.cell(2, 2, "기능 명세서 (Functional Specification)").font = font_title
    ws1.cell(3, 2, "PROJECT: Factory Operation AI Agent (제조·물류 공장 운영 및 설비 보전 질의응답 AI 에이전트)").font = font_desc

    ws1.cell(5, 2, "■ 문서 기본 정보").font = font_section

    doc_info = [
        ("문서 번호", "FS-FOA-001 (대응 URS: URS-FOA-001)"),
        ("문서명", "Factory Operation AI Agent 기능 명세서 (Functional Specification)"),
        ("버전", "v1.0 (최신 URS v0.1 반영판)"),
        ("작성일자", "2026-08-24"),
        ("작성자 / 부서", "송준환 / 제조AX솔루션팀"),
        ("승인자", "운영총괄 / QA팀"),
        ("문서 목적", "사용자 요구사항 정의서(URS v0.1)를 바탕으로 Factory Operation AI Agent의 22대 기능 요구사항(FR), 17대 핵심 세부 기능 명세(FN), Multi-SLM 엔진 처리 로직, 중계 서버 REST API 단일 인터페이스 및 Chat-First UI/UX 컴포넌트 인터랙션을 구체적으로 정의함"),
        ("참조 문서", "Factory Operation AI Agent URS v0.1 (urs_v0_1.docx), Concept.md v0.01, Features.md")
    ]

    for idx, (lbl, val) in enumerate(doc_info, start=6):
        c_lbl = ws1.cell(idx, 2, lbl)
        c_lbl.font = font_data_bold
        c_lbl.fill = fill_sub_header
        c_lbl.alignment = align_center
        c_lbl.border = border_cell

        c_val = ws1.cell(idx, 3, val)
        c_val.font = font_data
        c_val.alignment = align_left
        c_val.border = border_cell
        ws1.row_dimensions[idx].height = 24 if len(val) < 50 else 40

    ws1.cell(15, 2, "■ 문서 개정 이력").font = font_section

    rev_headers = ["버전", "개정일자", "주요 개정 내용 및 사유", "작성자", "승인자"]
    for c_idx, h in enumerate(rev_headers, start=2):
        cell = ws1.cell(16, c_idx, h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_header
    ws1.row_dimensions[16].height = 25

    rev_rows = [
        ("v0.00", "2026-03-20", "초기 기능 명세서 초안 작성", "자동화AX", "PM"),
        ("v0.50", "2026-06-10", "중계 서버 REST API 연동 및 Multi-SLM 파이프라인 구조 반영", "송준환", "PL"),
        ("v0.90", "2026-08-18", "RAG 지식 검색 및 설비 분석 기능 구체화", "송준환", "QA팀"),
        ("v1.0", "2026-08-24", "URS v0.1 정합성 전면 개정: 4대 핵심 가치(14개 FN) 및 플랫폼 기능(2개 FN), 22개 FR 매핑 체계화, 중계 서버 단일 통로(Query 전용) 및 Chat-First UI/UX 컴포넌트 명세 확정", "송준환", "운영총괄")
    ]

    for r_idx, r_data in enumerate(rev_rows, start=17):
        fill = fill_zebra if r_idx % 2 == 1 else fill_white
        for c_idx, val in enumerate(r_data, start=2):
            cell = ws1.cell(r_idx, c_idx, val)
            cell.font = font_data
            cell.fill = fill
            cell.alignment = align_center if c_idx in [2, 3, 5, 6] else align_left
            cell.border = border_cell
        ws1.row_dimensions[r_idx].height = 30 if len(r_data[2]) > 40 else 22

    ws1.column_dimensions['B'].width = 18.0
    ws1.column_dimensions['C'].width = 85.0
    ws1.column_dimensions['D'].width = 16.0
    ws1.column_dimensions['E'].width = 16.0
    ws1.column_dimensions['F'].width = 16.0

    # =========================================================================
    # SHEET 2: 2. 기능 목록표 (Function List / FR & FN Mapping)
    # =========================================================================
    ws2 = wb.create_sheet(title="2. 기능 목록표")
    ws2.views.sheetView[0].showGridLines = True

    ws2.cell(1, 1, "Factory Operation AI Agent - 전체 기능 목록표 (Function List & Traceability)").font = font_title
    ws2.row_dimensions[1].height = 30

    headers2 = [
        "기능 ID (FS-ID)", "대분류 (시스템 영역)", "중분류 (핵심 가치)", "소분류 (세부 기능그룹)",
        "기능명", "기능 요약 및 처리 내용", "접근 권한 (RBAC)", "우선순위", "대응 URS ID", "연계 모듈 / 인터페이스"
    ]
    for c_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(3, c_idx, h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_header
    ws2.row_dimensions[3].height = 28

    # 22 Requirements mapped with 17 Detailed Features and Core Enablers
    func_list = [
        # (1) 공통 대화 및 엔진 인터페이스
        ("FS-UI-01", "공통 인터랙션", "공통 UI/UX", "대화형 인터페이스", "대화형 멀티모달 채팅 UI & 프롬프트 처리", "자연어 텍스트, 음성(STT), 파일 첨부를 통한 통합 질의응답 및 피드백 표출", "All (Viewer+)", "M", "FR-001", "Chat Web/Tablet UI, STT Engine"),
        ("FS-ENG-01", "AI 코어 엔진", "Multi-SLM", "오케스트레이션", "Multi-SLM (Clas-Plan-Proc) 파이프라인 엔진", "질의 의도 분류(Clas), 실행 계획 수립(Plan), 지식/데이터 가공(Proc) 3단계 파이프라인", "All (Viewer+)", "M", "FR-002", "Clas-SLM / Plan-SLM / Proc-SLM"),
        
        # (2) [핵심 가치 1] 물류 설비 보전 지식 & SOP 매뉴얼 안내
        ("FS-SOP-01", "현장 보전 지능화", "1. 보전 지식 & SOP", "고장 조치 SOP", "물류 설비 고장 조치 SOP 단계별 가이드", "설비 알람 코드 발생 시 현장 표준 점검 및 1~N단계 조치 수순 자연어 안내", "User (Operator+)", "M", "FR-003 / FN-SOP-01", "Proc-SLM + Vector DB (SOP 임베딩)"),
        ("FS-SOP-02", "현장 보전 지능화", "1. 보전 지식 & SOP", "정비 매뉴얼 Q&A", "설비 부품/소모품 교체 및 정비 매뉴얼 Q&A", "정기 점검 주기, 체결 토크, 분해/조립 수순, 순정 부품 품번 매뉴얼 질의응답", "User (Operator+)", "M", "FR-004 / FN-SOP-02", "Proc-SLM + Vector DB (매뉴얼)"),
        ("FS-SOP-03", "현장 보전 지능화", "1. 보전 지식 & SOP", "트러블슈팅 탐색", "트러블슈팅 원인 규명 및 과거 조치 이력 탐색", "과거 동종 설비/알람 발생 사례 및 숙련 엔지니어의 실제 조치 이력 지식 탐색", "User (Operator+)", "M", "FR-005 / FN-SOP-03", "Vector DB (정비일지/조치이력)"),
        ("FS-SOP-04", "현장 보전 지능화", "1. 보전 지식 & SOP", "멀티모달 매뉴얼", "멀티모달 매뉴얼(PDF/도면) 실시간 파싱 & 질의", "현장 엔지니어가 업로드한 PDF 매뉴얼(최대 50MB) 실시간 파싱/청킹 및 도면 딥링크 질의", "User (Operator+)", "M", "FR-006 / FN-SOP-04", "PDF Parser Engine + In-Memory Vector"),

        # (3) [핵심 가치 2] 물류 설비 성능 및 가동률/알람 분석
        ("FS-EQP-01", "설비/물류 모니터링", "2. 설비 성능/알람 분석", "이송 설비 진단", "물류 이송 설비(AGV/무인지게차) 이상 진단 및 현황 질의", "AGV/무인지게차 실시간 노드 위치, 배터리 잔량, 주행/대기/이상 상태 표출", "All (Viewer+)", "M", "FR-007 / FN-EQP-01", "Relay REST API (GET /equip/agv)"),
        ("FS-EQP-02", "설비/물류 모니터링", "2. 설비 성능/알람 분석", "이송 시스템 성능", "자동화 이송 시스템 처리 성능(TPH) 및 병목 진단", "Sorting C/V, Lifter, Diverter, RGV, S/C 라인별 시간당 처리량 비교 및 병목 구간 식별", "All (Viewer+)", "M", "FR-008 / FN-EQP-02", "Relay REST API (GET /equip/conveyor)"),
        ("FS-EQP-03", "설비/물류 모니터링", "2. 설비 성능/알람 분석", "적재/투입 상태", "적재 및 투입 설비(로더/언로더) 상태 & 큐 정체 분석", "로더/언로더 작동 상태, 투입 대기량 및 버퍼 대기 큐 정체 요인 분석", "All (Viewer+)", "M", "FR-009 / FN-EQP-03", "Relay REST API (GET /equip/loader)"),
        ("FS-EQP-04", "설비/물류 모니터링", "2. 설비 성능/알람 분석", "가동률 분석", "설비 가동률(OEE) 및 비가동 원인 분석 리포트", "일/주/월간 설비별 시간가동률(OEE) 집계 및 비가동 원인 파레토 분석 인라인 표 렌더링", "User (Operator+)", "M", "FR-010 / FN-EQP-04", "Relay REST API (GET /kpi/oee)"),
        ("FS-EQP-05", "설비/물류 모니터링", "2. 설비 성능/알람 분석", "알람 통계 분석", "설비 알람 통계 및 빈발 장애 요약 분석", "설비/호기/알람코드별 발생 빈도 및 정지 시간 집계, 최다 빈발 장애 요약 브리핑", "User (Operator+)", "M", "FR-014 / FN-EQP-05", "Relay REST API (GET /alarm/stats)"),

        # (4) [핵심 가치 3] 자재 및 공정 재고 이상 추적
        ("FS-INV-01", "설비/물류 모니터링", "3. 재고 이상 추적", "창고 재고/FIFO", "자재 창고 재고 현황 & FIFO 유효기간 추적", "자재 창고 구역별 적재율, 유효기간(D-day) 카운트다운 및 선입선출(FIFO) 위반 경고", "All (Viewer+)", "M", "FR-011 / FN-INV-01", "Relay REST API (GET /inv/warehouse)"),
        ("FS-INV-02", "설비/물류 모니터링", "3. 재고 이상 추적", "버퍼 체류 분석", "공정 간 버퍼(WIP) 체류 시간 및 정체 분석", "공정 버퍼 적재량, 체류 시간(정상/주의/지연 인디케이터) 및 공용기 잔여량 분석", "All (Viewer+)", "M", "FR-012 / FN-INV-02", "Relay REST API (GET /inv/buffer)"),
        ("FS-INV-03", "설비/물류 모니터링", "3. 재고 이상 추적", "이송 지연 추적", "공정 간 물류 이송 지연 탐지 & 도착 예측", "공정 간 이동 중인 화물 이동 트래킹, 지연 병목 구간 탐지 및 도착 예상 시각 산출", "All (Viewer+)", "M", "FR-013 / FN-INV-03", "Relay REST API (GET /inv/transit)"),

        # (5) [핵심 가치 4] AI 시스템 운영 지원 산출물 자동 제공
        ("FS-ADV-01", "운영 지원 산출물", "4. 운영 지원 산출물", "PM 체크리스트", "표준 데이터 & 정기 PM 체크리스트 자동 생성", "신규/개조 설비 도입 시 태그 매핑 추천 및 일/주/월 정기 예방보전(PM) 체크리스트 생성", "User (Operator+)", "M", "FR-015 / FN-ADV-01", "AI Operations Advisor + Vector DB"),
        ("FS-ADV-02", "운영 지원 산출물", "4. 운영 지원 산출물", "시운전 가이드", "테스트 / 5단계 시운전 가이드 지원", "시스템 패치 또는 개조 시 단위/통합 테스트 케이스 및 5단계 시운전 절차서 초안 수립", "User (Operator+)", "M", "FR-016 / FN-ADV-02", "AI Operations Advisor + Knowledge"),
        ("FS-ADV-03", "운영 지원 산출물", "4. 운영 지원 산출물", "비상 대응(BCP)", "비상 대응(BCP) & 수동 전환 가이드", "상위 시스템 다운 시 현장 제어반 수동 운전(Manual Override) 전환 절차 및 핫라인 안내", "User (Operator+)", "M", "FR-017 / FN-ADV-03", "BCP Knowledge Module + Proc-SLM"),

        # (6) [플랫폼 기반 편의 기능] 개인화 프로젝트 워크스페이스
        ("FS-PWS-01", "지식 플랫폼", "5. 프로젝트 워크스페이스", "개인/팀 공간", "개인 / 팀별 프로젝트 워크스페이스", "개인/팀별 프로젝트 공간 생성, 전용 매뉴얼/도면 영구 보관 및 세션 초월 참조 지원", "User (Operator+)", "M", "FR-018 / FN-PWS-01", "Personalized Workspace DB + Storage"),
        ("FS-PWS-02", "지식 플랫폼", "5. 프로젝트 워크스페이스", "맞춤 정밀 검색", "프로젝트 전용 매뉴얼 맞춤 검색 (Scoped RAG)", "전체 공장 DB 대신 현재 활성화된 프로젝트 워크스페이스 문서로 검색 범위 한정(Namespace)", "User (Operator+)", "M", "FR-019 / FN-PWS-02", "Scoped Vector Namespace Engine"),

        # (7) 공통 UI/UX 컴포넌트 및 환각 차단 그라운딩
        ("FS-UI-02", "공통 인터랙션", "공통 UI/UX", "선택 칩", "대화형 선택 칩 (Option Chip UI)", "프롬프트 필수 인자(설비명/알람코드) 누락 시 에러 대신 클릭형 버튼 칩으로 즉시 제시", "All (Viewer+)", "M", "FR-020", "Interactive Option Chip Component"),
        ("FS-UI-03", "공통 인터랙션", "공통 UI/UX", "출처 인라인 배너", "원문 출처 인라인 배너 & 딥링크", "지식 검색 근거 문서명, 페이지 번호 표출 및 클릭 시 문서 뷰어 원문 위치 즉시 이동", "All (Viewer+)", "M", "FR-021", "Grounding Citation Banner"),
        ("FS-ENG-02", "AI 코어 엔진", "환각 억제/Grounding", "창작 차단", "환각 차단 및 답변 불가 안내 (Strict Grounding)", "지식 검색 유사도 미달(0.70 미만) 시 임의 추정 생성을 원천 차단하고 수동 확인 가이드 표출", "All (Viewer+)", "M", "FR-022", "Strict Grounding & Fallback Router")
    ]

    for r_idx, row in enumerate(func_list, start=4):
        fill = fill_zebra if r_idx % 2 == 1 else fill_white
        for c_idx, val in enumerate(row, start=1):
            cell = ws2.cell(r_idx, c_idx, val)
            cell.font = font_data_bold if c_idx in [1, 5, 8] else font_data
            cell.fill = fill
            cell.alignment = align_center if c_idx in [1, 2, 3, 4, 7, 8, 9] else align_left
            cell.border = border_cell
        ws2.row_dimensions[r_idx].height = 24

    ws2.column_dimensions['A'].width = 14.0
    ws2.column_dimensions['B'].width = 18.0
    ws2.column_dimensions['C'].width = 24.0
    ws2.column_dimensions['D'].width = 20.0
    ws2.column_dimensions['E'].width = 38.0
    ws2.column_dimensions['F'].width = 56.0
    ws2.column_dimensions['G'].width = 16.0
    ws2.column_dimensions['H'].width = 10.0
    ws2.column_dimensions['I'].width = 18.0
    ws2.column_dimensions['J'].width = 32.0

    # =========================================================================
    # SHEET 3: 3. 세부 기능 명세 (Detailed Specs)
    # =========================================================================
    ws3 = wb.create_sheet(title="3. 세부 기능 명세 (Detailed Specs)")
    ws3.views.sheetView[0].showGridLines = True

    ws3.cell(1, 1, "Factory Operation AI Agent - 17대 핵심 세부 기능 명세 (Detailed Specifications)").font = font_title
    ws3.row_dimensions[1].height = 30

    headers3 = [
        "기능 ID", "세부 기능 ID / 기능명", "기능 정의 및 개요 (Description)", "사전 조건 (Pre-condition)",
        "상세 처리 로직 및 알고리즘 (Processing Logic)", "입력 파라미터 규격 (Input)",
        "출력 파라미터 규격 (Output)", "연계 인터페이스 (Interface)", "예외 및 안전 처리 기준 (Fail-Safe)"
    ]
    for c_idx, h in enumerate(headers3, start=1):
        cell = ws3.cell(3, c_idx, h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_header
    ws3.row_dimensions[3].height = 28

    detailed_specs = [
        (
            "FS-SOP-01",
            "FN-SOP-01\n물류 설비 이상 조치 SOP 1~N단계 수순 가이드",
            "물류 설비(AGV, RGV, S/C 등) 알람 코드 발생 시 현장 표준 점검 및 조치 수순(SOP)을 단계별 자연어로 안내하고 관련 도면 링크 제공",
            "사용자 로그인 완료, SOP 지식 Vector DB 정상 가동",
            "1. 프롬프트 수신\n➔ 2. Clas-SLM이 알람코드/설비명 식별\n➔ 3. Plan-SLM이 SOP 지식 검색 계획 수립\n➔ 4. Proc-SLM이 Vector DB 검색(유사도 >= 0.70)\n➔ 5. 1~N단계 수순 요약 및 안전 확인 수칙, 출처(PDF P.XX) 링크 생성",
            "• prompt: string (필수)\n• equip_id: string (선택)\n• alarm_code: string (선택)",
            "• sop_steps: array of string\n• safety_warning: string\n• citations: array of {doc_id, page, score}",
            "내부 Vector DB (SOP 임베딩)",
            "검색 유사도 < 0.70 미만 시 추정 답변 차단, 현장 수동 점검 안내문 표출(FS-ENG-02)"
        ),
        (
            "FS-SOP-02",
            "FN-SOP-02\n설비 부품/소모품 교체 및 정비 매뉴얼 Q&A",
            "정기 점검 주기, 소모품/부품 교체 가이드, 체결 토크 기준, 분해/조립 수순 및 순정 부품 품번을 매뉴얼 기반으로 질의응답",
            "정비 매뉴얼 지식 베이스 구축 완료",
            "1. 부품/점검 질의 수신\n➔ 2. 부품명/설비 모델 추출\n➔ 3. 매뉴얼 Vector DB 탐색\n➔ 4. 교체 수순, 권장 토크, 순정 품번 및 주의사항 추출\n➔ 5. 대화형 카드 형태로 정렬 표출",
            "• equip_model: string (필수)\n• part_name: string (선택)\n• query_topic: enum (TORQUE, CYCLE, PART_NO)",
            "• part_no: string\n• spec_torque: string\n• replace_steps: array of string\n• caution: string",
            "내부 Vector DB (정비 매뉴얼)",
            "부품 규격 불일치 시 오적용 방지 경고 및 상위 모델 매뉴얼 확인 권고"
        ),
        (
            "FS-SOP-03",
            "FN-SOP-03\n트러블슈팅 원인 규명 및 과거 조치 이력 탐색",
            "과거 동종 설비/알람 발생 사례 및 보전 엔지니어의 실제 조치 이력(정비 일지)을 검색하여 고장 원인 분석 및 해결책 추천",
            "정비 일지 및 트러블슈팅 이력 DB 인덱싱 완료",
            "1. 고장 증상 질의\n➔ 2. 증상 키워드 및 설비 코드 기반 하이브리드 검색\n➔ 3. 과거 조치 이력 Top 3 추출\n➔ 4. 원인별 조치 성공률 및 해결 수순 브리핑",
            "• symptom_desc: string (필수)\n• equip_type: string (선택)",
            "• trouble_cases: array of {case_id, root_cause, action_taken, solved_date}",
            "내부 Vector DB (정비일지 이력)",
            "이력 데이터 부족 시 표준 SOP 우선 안내로 자동 전환"
        ),
        (
            "FS-SOP-04",
            "FN-SOP-04\n멀티모달 매뉴얼(PDF/도면) 파싱 및 지식 검색",
            "사용자가 업로드한 신규 PDF 정비 매뉴얼(최대 50MB)을 실시간 파싱하여 문서 내 세부 조치법 검색 및 도면 딥링크 제공",
            "파일 크기 50MB 이하, 지원 포맷(PDF, DOCX, PPTX, XLSX, TXT, 이미지)",
            "1. 파일 드래그&드롭 업로드\n➔ 2. 용량 및 확장자 검증\n➔ 3. PDF Parser가 텍스트/도표 청킹\n➔ 4. Vector 임베딩 메모리 인덱싱\n➔ 5. 질의 키워드 기반 RAG 응답 및 원문 링크 표출",
            "• file: binary (최대 50MB)\n• query: string (필수)",
            "• summary_answer: string\n• ref_pages: array of int\n• drawing_links: array of string",
            "PDF Parser Engine + In-Memory Vector",
            "50MB 초과 또는 미지원 포맷 첨부 시 즉시 업로드 차단 팝업 표출"
        ),
        (
            "FS-EQP-01",
            "FN-EQP-01\n물류 이송 설비(AGV/무인지게차) 이상 진단 및 현황 질의",
            "전체 또는 특정 AGV/무인지게차의 현재 노드 위치, 배터리 잔량, 주행/대기/이상 상태를 중계 서버 경유 실시간 표출",
            "중계 서버 REST API 통신 정상 연결",
            "1. AGV 상태 질의 접수\n➔ 2. Clas-SLM 의도 판별\n➔ 3. Proc-SLM이 중계 서버 API(GET /api/v1/relay/equip/agv) 호출\n➔ 4. JSON 응답 수신\n➔ 5. 배터리 부족/이상 설비 경고 하이라이트 동적 표 렌더링",
            "• query: string (필수)\n• agv_id: string (선택)\n• zone_id: string (선택)",
            "• agv_list: array of {agv_id, status, node_loc, battery_pct, current_job, alarm_code}",
            "중계 서버 REST API (`GET /api/v1/relay/equip/agv`)",
            "중계 서버 1.5초 지연 경고, 3초 초과 시 캐시 상태 표시 및 통신 오류 안내"
        ),
        (
            "FS-EQP-02",
            "FN-EQP-02\n자동화 이송 시스템 처리 성능 및 병목 진단",
            "Sorting C/V, Lifter, Diverter, RGV, S/C 라인별 시간당 처리량(TPH), 사이클 타임 비교 및 병목 구간 식별 진단",
            "중계 서버 REST API 연동 정상",
            "1. 성능/병목 질의 접수\n➔ 2. 중계 서버 API(GET /api/v1/relay/equip/conveyor) 호출\n➔ 3. 라인별 실적 TPH vs 목표 TPH 비교 연산\n➔ 4. 병목 1순위 설비 및 정체 구간 도출\n➔ 5. 병목 진단 카드 출력",
            "• line_id: string (선택)\n• equip_type: enum (CONVEYOR, LIFTER, RGV, SC)",
            "• tph_summary: array of {line_id, target_tph, actual_tph, bottleneck_flag}\n• diag_msg: string",
            "중계 서버 REST API (`GET /api/v1/relay/equip/conveyor`)",
            "라인 데이터 결측 시 최근 1시간 평균값 기반 추세 정보 대체 표출"
        ),
        (
            "FS-EQP-03",
            "FN-EQP-03\n적재 및 투입 설비 상태 및 큐 정체 분석",
            "투입 로더/언로더 작동 상태, 투입 대기량 및 버퍼 대기 큐 정체 요인을 분석하여 피드백",
            "중계 서버 연동 정상",
            "1. 투입/적재 현황 질의\n➔ 2. 중계 서버 API(GET /api/v1/relay/equip/loader) 호출\n➔ 3. 로더별 작동 모드 및 대기 큐 적재량 분석\n➔ 4. 투입 지연 요인 요약 카드 렌더링",
            "• loader_id: string (선택)\n• zone_id: string (선택)",
            "• loader_status: array of {loader_id, state, queue_cnt, max_capacity, dwell_sec}",
            "중계 서버 REST API (`GET /api/v1/relay/equip/loader`)",
            "큐 적재율 90% 초과 시 '만차 임박 경고' 인라인 알림 표출"
        ),
        (
            "FS-EQP-04",
            "FN-EQP-04\n설비 가동률(OEE) 및 비가동 원인 분석 리포트",
            "일간/주간/월간 설비별 시간가동률(OEE) 집계 및 비가동 원인 파레토 분석 리포트를 대화창 내 인라인 동적 표로 제공",
            "중계 서버 가동률 데이터 집계 완료",
            "1. 가동률/비가동 질의\n➔ 2. 중계 서버 API(GET /api/v1/relay/kpi/oee) 호출\n➔ 3. 기간별 가동률(%) 및 비가동 사유(설비고장/대기/작업지연) 집계\n➔ 4. 파레토 비율 연산\n➔ 5. 인라인 마크다운 표 및 요약 브리핑",
            "• period: enum (DAILY, WEEKLY, MONTHLY)\n• equip_group: string (선택)",
            "• oee_kpi: {target_oee, actual_oee, availability}\n• downtime_reasons: array of {reason, hours, ratio_pct}",
            "중계 서버 REST API (`GET /api/v1/relay/kpi/oee`)",
            "집계 기간 미지정 시 당일(오늘) 누적 실적 기본 적용"
        ),
        (
            "FS-EQP-05",
            "FN-EQP-05\n설비 알람 통계 및 빈발 장애 요약 분석",
            "설비/호기/알람코드별 발생 빈도 및 정지 시간 통계 분석을 통해 최다 빈발 알람 요약 브리핑 제공",
            "알람 통계 데이터 연동 정상",
            "1. 알람 통계 질의\n➔ 2. 중계 서버 API(GET /api/v1/relay/alarm/stats) 호출\n➔ 3. 발생 횟수 및 누적 정지 시간 기준 내림차순 정렬\n➔ 4. 최빈발 Top 3 알람 및 권장 예방보전 조치 안내",
            "• date_range: string (선택)\n• top_n: int (기본 5)",
            "• alarm_ranks: array of {rank, alarm_code, equip_id, occur_count, stop_mins, main_cause}",
            "중계 서버 REST API (`GET /api/v1/relay/alarm/stats`)",
            "통계 데이터 0건 시 '조회 기간 내 특이 알람 미발생' 상태 안내"
        ),
        (
            "FS-INV-01",
            "FN-INV-01\n자재 창고 재고 현황 & FIFO 유효기간 추적",
            "자재 창고 구역별 적재율, 유효기간(D-day) 카운트다운 관리 및 선입선출(FIFO) 위반 경고를 인라인 표로 표출",
            "WMS/MES 연동 중계 서버 정상 가동",
            "1. 재고/FIFO 질의\n➔ 2. 중계 서버 API(GET /api/v1/relay/inv/warehouse) 호출\n➔ 3. D-day 임박 품목 및 FIFO 역전 품목 탐지\n➔ 4. 위험도별 색상(적/황/녹) 태그 부여 동적 표 출력",
            "• zone_id: string (선택)\n• d_day_limit: int (기본 3일)",
            "• warehouse_status: {capacity_pct, total_lots}\n• warning_items: array of {lot_no, item, d_day, fifo_violation}",
            "중계 서버 REST API (`GET /api/v1/relay/inv/warehouse`)",
            "FIFO 위반 품목 탐지 시 최상단 고정 배치 및 관리자 알림 권고"
        ),
        (
            "FS-INV-02",
            "FN-INV-02\n공정 간 버퍼(WIP) 체류 시간 및 정체 분석",
            "공정 간 버퍼 구역의 재공품 체류 시간 실시간 분석 및 임계치(45분) 초과 시 지연 경고 인디케이터 표출",
            "MES 연동 중계 서버 정상 가동",
            "1. 버퍼 정체 질의\n➔ 2. 중계 서버 API(GET /api/v1/relay/inv/buffer) 호출\n➔ 3. 체류 시간 45분 초과 WIP 품목 필터링\n➔ 4. 체류 상태(정상:녹 / 주의:황 / 지연:적) 인디케이터 바인딩 및 표 표출",
            "• buffer_zone_id: string (선택)\n• threshold_mins: int (기본 45분)",
            "• wip_items: array of {pallet_id, item_name, dwell_mins, status_tag, loc_rack}",
            "중계 서버 REST API (`GET /api/v1/relay/inv/buffer`)",
            "MES 조회 지연 시 '데이터 갱신 지연(최근 수집 시각: XX)' 배너 표출"
        ),
        (
            "FS-INV-03",
            "FN-INV-03\n공정 간 물류 이송 지연 탐지 & 도착 예측",
            "공정 간 이동 중인 화물의 운송 상태를 추적하고 이송 지연 병목 구간 탐지 및 도착 예상 시각 산출",
            "물류 관제 연동 중계 서버 정상 가동",
            "1. 이송 현황 질의\n➔ 2. 중계 서버 API(GET /api/v1/relay/inv/transit) 호출\n➔ 3. 표준 이송 시간 대비 지연 화물 탐지\n➔ 4. 잔여 거리/속도 기반 도착 예상 시각 계산\n➔ 5. 이송 진단 카드 표출",
            "• from_process: string (선택)\n• to_process: string (선택)",
            "• transit_cards: array of {cargo_id, transport_equip, eta, delay_mins, status}",
            "중계 서버 REST API (`GET /api/v1/relay/inv/transit`)",
            "이송 정지 감지 시 즉시 '정체 구간 발생' 경고 카드 표출"
        ),
        (
            "FS-ADV-01",
            "FN-ADV-01\n표준 데이터 & 정기 PM 체크리스트 생성",
            "신규/개조 설비 도입 시 태그 매핑 표준 추천 및 일/주/월 정기 예방보전(PM) 체크리스트 양식 자동 작성",
            "설비 모델명 또는 매뉴얼 사양 등록 완료",
            "1. PM 체크리스트 요청\n➔ 2. AI Operations Advisor 가동\n➔ 3. 유사 설비 표준 지식 베이스 검색\n➔ 4. 부위별(모터/센서/체결부/배선) 점검 항목 및 주기(일/주/월) 표 생성\n➔ 5. 마크다운/엑셀 서식 출력",
            "• equip_model: string (필수)\n• pm_cycle: enum (DAILY, WEEKLY, MONTHLY)",
            "• pm_checklist_table: array of {item_no, part, check_method, criteria, cycle}",
            "AI Operations Advisor + Vector DB",
            "참조 표준 결측 시 산업 표준 안전 점검 기본 템플릿 대체 적용"
        ),
        (
            "FS-ADV-02",
            "FN-ADV-02\n테스트 / 5단계 시운전 가이드 지원",
            "시스템 패치 또는 설비 개조 시 단위/통합 테스트 케이스 및 5단계 시운전 절차서 초안 자동 수립",
            "개조 사양 또는 변경 로직 정보 입력 완료",
            "1. 시운전 가이드 요청\n➔ 2. 수정 사양 및 인터록 영향 분석\n➔ 3. 5단계 시운전 수순(단동 ➔ 무부하 ➔ 부하 ➔ 인터록 ➔ 연동) 절차서 생성\n➔ 4. 테스트 시나리오/사전조건/기대결과 표 렌더링",
            "• system_change_desc: string (필수)\n• target_equip: string (필수)",
            "• commissioning_guide: {steps: array of string, test_cases: array of {id, pre, test, expect}}",
            "AI Operations Advisor + Knowledge Engine",
            "안전 인터록 테스트 항목 최우선 필수 항목으로 자동 삽입"
        ),
        (
            "FS-ADV-03",
            "FN-ADV-03\n비상 대응(BCP) & 수동 전환 가이드",
            "상위 시스템 다운 또는 통신 장애 시 현장 제어반 수동 운전(Manual Override) 전환 절차, 점검 항목 및 비상 연락망 안내",
            "BCP 지식 모듈 탑재 완료",
            "1. 비상 수동 전환 질의\n➔ 2. 장애 대상 시스템 식별\n➔ 3. BCP SOP 표준 절차서 탐색\n➔ 4. 수동 운전 3단계 수순, 안전 확인 체크리스트 및 비상 핫라인 연락망 출력",
            "• emergency_scenario: string (필수)\n• location: string (선택)",
            "• bcp_procedure: array of string\n• safety_checks: array of string\n• hotline: array of {role, contact}",
            "BCP Knowledge Module + Proc-SLM",
            "비상 상황 시 핵심 수순 상단 요약 강조 및 오프라인 인쇄 양식 지원"
        ),
        (
            "FS-PWS-01",
            "FN-PWS-01\n개인 / 팀별 프로젝트 워크스페이스",
            "개인/작업반 단위로 프로젝트 공간을 생성하고 자주 참조하는 매뉴얼(PDF), 도면을 영구 보관하여 세션 초월 재참조 지원",
            "사용자 계정 인증 완료",
            "1. 워크스페이스 생성 명령\n➔ 2. 식별자(ID) 부여\n➔ 3. 매뉴얼/도면 파일 업로드\n➔ 4. 전용 Vector DB 파싱 및 독립 Namespace 인덱싱\n➔ 5. 프로젝트 지식 패널 관리",
            "• workspace_name: string (필수)\n• file_list: array of file",
            "• workspace_id: string\n• doc_status: array of {doc_name, size, chunk_cnt}",
            "Personalized Workspace DB + Storage",
            "프로젝트당 용량 한도(500MB) 초과 시 용량 정리 안내"
        ),
        (
            "FS-PWS-02",
            "FN-PWS-02\n프로젝트 전용 매뉴얼 맞춤 검색",
            "전체 공장 DB 대신 현재 활성화된 프로젝트 워크스페이스 내 문서로만 검색 범위를 한정(Scoped RAG)하여 타 설비 혼선 차단",
            "활성화된 프로젝트 워크스페이스 선택 완료",
            "1. 워크스페이스 활성화 토글\n➔ 2. 질의 프롬프트 접수\n➔ 3. RAG 탐색 범위를 해당 워크스페이스 Namespace로 엄격 제한\n➔ 4. 전용 문서 기반 맞춤형 정밀 답변 및 출처 인덱스 도출",
            "• active_workspace_id: string (필수)\n• query: string (필수)",
            "• scoped_answer: string\n• doc_citations: array of {doc_name, page, score}",
            "Scoped Vector Namespace Engine",
            "지정 워크스페이스 내 관련 문서 부재 시 '프로젝트 내 관련 문서 없음' 명시 후 전체 검색 전환 여부 질의"
        )
    ]

    for r_idx, row in enumerate(detailed_specs, start=4):
        fill = fill_zebra if r_idx % 2 == 1 else fill_white
        for c_idx, val in enumerate(row, start=1):
            cell = ws3.cell(r_idx, c_idx, val)
            cell.font = font_code if c_idx in [1, 6, 7] else (font_data_bold if c_idx == 2 else font_data)
            cell.fill = fill
            cell.alignment = align_top_center if c_idx == 1 else align_top_left
            cell.border = border_cell
        ws3.row_dimensions[r_idx].height = 95

    ws3.column_dimensions['A'].width = 14.0
    ws3.column_dimensions['B'].width = 28.0
    ws3.column_dimensions['C'].width = 34.0
    ws3.column_dimensions['D'].width = 26.0
    ws3.column_dimensions['E'].width = 46.0
    ws3.column_dimensions['F'].width = 34.0
    ws3.column_dimensions['G'].width = 34.0
    ws3.column_dimensions['H'].width = 28.0
    ws3.column_dimensions['I'].width = 30.0

    # =========================================================================
    # SHEET 4: 4. 인터페이스 및 API 명세
    # =========================================================================
    ws4 = wb.create_sheet(title="4. 인터페이스 및 API 명세")
    ws4.views.sheetView[0].showGridLines = True

    ws4.cell(1, 1, "Factory Operation AI Agent - 중계 서버 REST API 인터페이스 명세 (Relay API Specs)").font = font_title
    ws4.row_dimensions[1].height = 30

    headers4 = [
        "인터페이스 ID", "인터페이스명", "송신 (Source)", "수신 (Target)",
        "통신 방식 (Method / URI)", "Request Body / Parameters", "Response Body 규격",
        "목표 응답시간", "안전 및 에러 처리 기준 (Fail-Safe)"
    ]
    for c_idx, h in enumerate(headers4, start=1):
        cell = ws4.cell(3, c_idx, h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_header
    ws4.row_dimensions[3].height = 28

    api_specs = [
        (
            "IF-RELAY-001",
            "AGV 및 무인지게차 위치/상태 조회 API",
            "AI 에이전트",
            "중계 서버 ➔ ACS/PLC",
            "GET /api/v1/relay/equip/agv",
            "Query Parameters:\n• zone_id: string (선택)\n• agv_id: string (선택)\n• status: enum (RUN, IDLE, ERROR, ALL)",
            "{\n  \"status\": 200,\n  \"data\": [\n    {\n      \"agv_id\": \"AGV-03\",\n      \"loc\": \"N-104\",\n      \"battery\": 78,\n      \"state\": \"RUN\",\n      \"current_job\": \"JOB-8821\"\n    }\n  ]\n}",
            "1.0초 이내",
            "통신 타임아웃(3초) 시 최근 캐시 데이터 표출 및 상단에 통신 지연 경고 배너 표출"
        ),
        (
            "IF-RELAY-002",
            "자동화 이송 설비(컨베이어/소터/RGV) 성능 조회 API",
            "AI 에이전트",
            "중계 서버 ➔ SCADA/WCS",
            "GET /api/v1/relay/equip/conveyor",
            "Query Parameters:\n• line_id: string (선택)\n• equip_type: enum (CV, LIFTER, RGV, SC)",
            "{\n  \"status\": 200,\n  \"data\": [\n    {\n      \"line_id\": \"LINE-01\",\n      \"target_tph\": 150,\n      \"actual_tph\": 138,\n      \"bottleneck_equip\": \"LIFTER-02\"\n    }\n  ]\n}",
            "1.5초 이내",
            "중계 서버 연동 실패 시 '이송 성능 데이터 수집 지연' 안내 표출"
        ),
        (
            "IF-RELAY-003",
            "적재/투입 로더 및 언로더 상태 조회 API",
            "AI 에이전트",
            "중계 서버 ➔ PLC/WCS",
            "GET /api/v1/relay/equip/loader",
            "Query Parameters:\n• loader_id: string (선택)\n• zone_id: string (선택)",
            "{\n  \"status\": 200,\n  \"data\": [\n    {\n      \"loader_id\": \"LDR-01\",\n      \"state\": \"RUN\",\n      \"queue_cnt\": 12,\n      \"max_cap\": 15\n    }\n  ]\n}",
            "1.5초 이내",
            "대기 큐 임계치 초과 시 경고 태그 부여"
        ),
        (
            "IF-RELAY-004",
            "설비 시간가동률(OEE) 및 비가동 분석 API",
            "AI 에이전트",
            "중계 서버 ➔ MES/SCADA",
            "GET /api/v1/relay/kpi/oee",
            "Query Parameters:\n• period: enum (DAILY, WEEKLY, MONTHLY)\n• equip_group: string (선택)",
            "{\n  \"status\": 200,\n  \"data\": {\n    \"oee_pct\": 92.4,\n    \"avail_pct\": 95.1,\n    \"downtime_reasons\": [\n      {\"reason\": \"센서오염\", \"mins\": 45, \"ratio\": 52.0}\n    ]\n  }\n}",
            "1.5초 이내",
            "비가동 사유 파레토 정렬 데이터 반환, 결측 시 최근 확정 실적 대체"
        ),
        (
            "IF-RELAY-005",
            "자재 창고 재고 및 FIFO/유효기간 조회 API",
            "AI 에이전트",
            "중계 서버 ➔ WMS",
            "GET /api/v1/relay/inv/warehouse",
            "Query Parameters:\n• zone_id: string (선택)\n• d_day_limit: int (기본 3)",
            "{\n  \"status\": 200,\n  \"data\": {\n    \"capacity_pct\": 84.5,\n    \"warning_lots\": [\n      {\"lot_no\": \"LOT-2026-A\", \"item\": \"원자재A\", \"d_day\": 2, \"fifo_violation\": true}\n    ]\n  }\n}",
            "1.5초 이내",
            "FIFO 위반 품목 최상단 정렬 반환, DB 과부하 방지 페이징(Page Size 50)"
        ),
        (
            "IF-RELAY-006",
            "공정 버퍼(WIP) 체류 시간 및 정체 조회 API",
            "AI 에이전트",
            "중계 서버 ➔ MES/WMS",
            "GET /api/v1/relay/inv/buffer",
            "Query Parameters:\n• buffer_zone_id: string (선택)\n• over_mins: int (기본 45)",
            "{\n  \"status\": 200,\n  \"data\": [\n    {\n      \"pallet_id\": \"P-9021\",\n      \"item\": \"WIP-A\",\n      \"dwell_mins\": 58,\n      \"status\": \"DELAY\"\n    }\n  ]\n}",
            "1.5초 이내",
            "체류 시간 45분 초과 품목에 대해 DELAY 플래그 자동 지정"
        ),
        (
            "IF-RELAY-007",
            "공정 간 물류 이송 현황 및 도착 예정 조회 API",
            "AI 에이전트",
            "중계 서버 ➔ WCS/ACS",
            "GET /api/v1/relay/inv/transit",
            "Query Parameters:\n• from_process: string (선택)\n• to_process: string (선택)",
            "{\n  \"status\": 200,\n  \"data\": [\n    {\n      \"cargo_id\": \"C-104\",\n      \"transport_agv\": \"AGV-05\",\n      \"eta\": \"14:25:00\",\n      \"delay_mins\": 8\n    }\n  ]\n}",
            "1.5초 이내",
            "이송 정체 구간 발생 시 bottleneck 경고 포함"
        ),
        (
            "IF-RELAY-008",
            "설비 알람 발생 이력 및 빈발 통계 조회 API",
            "AI 에이전트",
            "중계 서버 ➔ SCADA/DB",
            "GET /api/v1/relay/alarm/stats",
            "Query Parameters:\n• top_n: int (기본 5)\n• date_range: string (선택)",
            "{\n  \"status\": 200,\n  \"data\": [\n    {\n      \"alarm_code\": \"ERR-104\",\n      \"equip_type\": \"AGV\",\n      \"occur_count\": 14,\n      \"total_stop_mins\": 68\n    }\n  ]\n}",
            "1.5초 이내",
            "발생 횟수 내림차순 정렬 반환"
        ),
        (
            "IF-VEC-001",
            "사내 설비 보전 SOP 및 정비 매뉴얼 지식 검색 API",
            "AI 에이전트 (Proc-SLM)",
            "내부 Vector DB",
            "Internal Vector Search (Embedding Similarity)",
            "{\n  \"query_vector\": [0.012, -0.045, ...],\n  \"top_k\": 3,\n  \"threshold\": 0.70,\n  \"namespace\": \"default or workspace_id\"\n}",
            "{\n  \"results\": [\n    {\n      \"doc_id\": \"SOP-AGV-001\",\n      \"page\": 12,\n      \"text\": \"ERR-104 센서 점검...\",\n      \"score\": 0.88\n    }\n  ]\n}",
            "0.5초 이내",
            "유사도 score < 0.70 미달 시 빈 결과 반환 ➔ Proc-SLM 환각 차단 가동(FS-ENG-02)"
        )
    ]

    for r_idx, row in enumerate(api_specs, start=4):
        fill = fill_zebra if r_idx % 2 == 1 else fill_white
        for c_idx, val in enumerate(row, start=1):
            cell = ws4.cell(r_idx, c_idx, val)
            cell.font = font_code if c_idx in [1, 5, 6, 7] else (font_data_bold if c_idx == 2 else font_data)
            cell.fill = fill
            cell.alignment = align_top_center if c_idx in [1, 3, 4, 8] else align_top_left
            cell.border = border_cell
        ws4.row_dimensions[r_idx].height = 110

    ws4.column_dimensions['A'].width = 16.0
    ws4.column_dimensions['B'].width = 28.0
    ws4.column_dimensions['C'].width = 16.0
    ws4.column_dimensions['D'].width = 22.0
    ws4.column_dimensions['E'].width = 30.0
    ws4.column_dimensions['F'].width = 40.0
    ws4.column_dimensions['G'].width = 44.0
    ws4.column_dimensions['H'].width = 16.0
    ws4.column_dimensions['I'].width = 32.0

    # =========================================================================
    # SHEET 5: 5. 화면 및 인터랙션 명세
    # =========================================================================
    ws5 = wb.create_sheet(title="5. 화면 및 인터랙션 명세")
    ws5.views.sheetView[0].showGridLines = True

    ws5.cell(1, 1, "Factory Operation AI Agent - 화면 및 UI/UX 인터랙션 명세 (UI/UX Specs)").font = font_title
    ws5.row_dimensions[1].height = 30

    headers5 = [
        "화면 / 컴포넌트 ID", "컴포넌트명", "위치 (Layout)", "주요 UI 요소",
        "사용자 인터랙션 이벤트 (Action)", "시스템 반응 및 동적 렌더링 동작 (Response)", "비고 / 예외"
    ]
    for c_idx, h in enumerate(headers5, start=1):
        cell = ws5.cell(3, c_idx, h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_header
    ws5.row_dimensions[3].height = 28

    ui_specs = [
        (
            "UI-CHAT-001",
            "대화형 채팅 입력창 (Chat Input Bar)",
            "화면 하단 고정 바",
            "자연어 텍스트 입력창, 음성(STT) 마이크 아이콘, 파일 첨부 클립, 전송(Send) 버튼",
            "• 텍스트 입력 후 Enter 또는 전송 클릭\n• 마이크 클릭 후 음성 발화",
            "• 질의 메시지 우측 버블 즉시 표출\n• 좌측에 Multi-SLM 생각 중(Spinner) 인디케이터 즉시 가동\n• 답변 수신 시 스트리밍 렌더링",
            "빈 메시지 전송 차단, 1000자 초과 시 입력 제한 안내"
        ),
        (
            "UI-CHIP-002",
            "대화형 선택 칩 (Option Chip UI)",
            "대화 메시지 하단",
            "설비명/알람코드/질의선택 버튼 칩 (예: [AGV 3호기], [S/C 1호기], [가동률 분석])",
            "인자 누락 시 제시된 버튼 칩 클릭",
            "• 클릭된 칩의 텍스트가 질의 컨텍스트에 자동 바인딩\n• 즉시 후속 파이프라인(RAG/API) 자동 실행 (키보드 조작 최소화)",
            "최대 5개까지 가로 스크롤 칩 표출, 1회 클릭 후 비활성화"
        ),
        (
            "UI-UPLD-003",
            "멀티모달 드래그&드롭 존 (Upload Zone)",
            "채팅 모달 / 첨부 팝업",
            "점선 드롭 존, 첨부 파일명, 용량 프로그레스 바, 파싱 상태 인디케이터, 삭제(X) 버튼",
            "PC/태블릿에서 파일 드래그 앤 드롭 또는 파일 첨부 클릭",
            "• 클라이언트 50MB 용량 및 확장자(PDF/Office/이미지) 1차 검증\n• 실시간 파싱 진행률(0~100%) 표시 후 In-Memory RAG 준비 완료 배너 표시",
            "50MB 초과 또는 실행 파일(.exe 등) 첨부 시 즉시 업로드 차단 팝업"
        ),
        (
            "UI-CARD-004",
            "대화형 Markdown 동적 표 및 요약 카드",
            "대화창 내 답변 영역",
            "Markdown Table 그리드, KPI 요약 배지(OEE%), 지연 경고 태그(정상:녹/주의:황/지연:적), 정렬/필터",
            "• 표 내 정렬 화살표 클릭\n• 요약 카드 접기/펼치기 토글",
            "• 가동률, WIP 체류 목록, 알람 통계, PM 체크리스트를 정형화된 동적 표로 인라인 렌더링\n• 뷰어/대시보드 전환 없이 대화창 내 즉시 확인",
            "모바일/태블릿 가로 스크롤 반응형 CSS 적용"
        ),
        (
            "UI-CITE-005",
            "원문 출처 인라인 배너 & 딥링크",
            "답변 버블 최하단",
            "출처 문서명 배지, 참조 페이지 번호 (예: [📄 AGV 정비매뉴얼.pdf P.45]), 딥링크 버튼",
            "출처 배너 또는 페이지 번호 클릭",
            "• 사내 문서 뷰어가 팝업 또는 분할 뷰로 열리며 해당 페이지/도면 하이라이트 위치로 즉시 스크롤 이동",
            "지식 DB 참조 질의(SOP/매뉴얼)에만 표출, API 단독 조회 시 미표출"
        ),
        (
            "UI-PWS-006",
            "프로젝트 워크스페이스 관리 패널",
            "화면 좌측 사이드바",
            "프로젝트 워크스페이스 목록, 신규 생성 버튼, 등록 문서 목록(PDF/도면), 활성화 ON/OFF 토글",
            "• 프로젝트 클릭하여 활성화\n• 신규 문서 업로드 및 삭제",
            "• 활성화 시 대화창 상단에 '[AGV 3호기 전용 공간] 적용 중' 상태 배너 표출\n• RAG 질의 시 해당 워크스페이스 전용 Namespace로 검색 스코프 즉시 한정",
            "프로젝트 미선택 시 기본 공장 전체 지식 베이스 검색"
        ),
        (
            "UI-WARN-007",
            "안전 및 검색 불가 안내 인라인 알림 (Grounding Fallback)",
            "대화창 내 피드백 영역",
            "경고 아이콘(⚠️), '관련 보전 SOP 근거를 찾을 수 없습니다' 안내문, 현장 수동 확인 권고",
            "안내문 하단 [관제실 연락망 보기] 또는 [전체 검색 재시도] 클릭",
            "• 지식 유사도 < 0.70 미달 시 환각 답변을 차단하고 규격화된 안내문 표출\n• 수동 관제실 핫라인 정보 즉시 제시",
            "임의 텍스트 생성 엄격 차단"
        )
    ]

    for r_idx, row in enumerate(ui_specs, start=4):
        fill = fill_zebra if r_idx % 2 == 1 else fill_white
        for c_idx, val in enumerate(row, start=1):
            cell = ws5.cell(r_idx, c_idx, val)
            cell.font = font_code if c_idx == 1 else (font_data_bold if c_idx == 2 else font_data)
            cell.fill = fill
            cell.alignment = align_top_center if c_idx in [1, 3] else align_top_left
            cell.border = border_cell
        ws5.row_dimensions[r_idx].height = 70

    ws5.column_dimensions['A'].width = 16.0
    ws5.column_dimensions['B'].width = 28.0
    ws5.column_dimensions['C'].width = 18.0
    ws5.column_dimensions['D'].width = 34.0
    ws5.column_dimensions['E'].width = 30.0
    ws5.column_dimensions['F'].width = 46.0
    ws5.column_dimensions['G'].width = 28.0

    # Save to Functional_Specification_v1.0.xlsx
    wb.save("Functional_Specification_v1.0.xlsx")
    print("Successfully updated Functional_Specification_v1.0.xlsx!")

if __name__ == "__main__":
    build_functional_specification()
