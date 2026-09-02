import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_ia_excel():
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # Styles Definition
    # ---------------------------------------------------------
    font_family = "Malgun Gothic"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F2937")
    subtitle_font = Font(name=font_family, size=11, color="4B5563")
    section_font = Font(name=font_family, size=12, bold=True, color="1E3A8A")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    cell_font = Font(name=font_family, size=9.5, color="1F2937")
    cell_font_bold = Font(name=font_family, size=9.5, bold=True, color="1F2937")
    cell_font_mono = Font(name="Consolas", size=9.5, color="1F2937")
    note_font = Font(name=font_family, size=9, color="6B7280", italic=True)
    badge_font = Font(name=font_family, size=9, bold=True, color="1E40AF")
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    category_a_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid") # 연녹색
    category_b_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # 연파랑
    category_c_fill = PatternFill(start_color="FAF5FF", end_color="FAF5FF", fill_type="solid") # 연보라
    category_d_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid") # 연노랑
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    highlight_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D1D5DB")
    header_border_side = Side(border_style="thin", color="1E3A8A")
    
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=header_border_side, right=header_border_side, top=header_border_side, bottom=header_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # ---------------------------------------------------------
    # Sheet 1: 개요 및 표지 (Overview)
    # ---------------------------------------------------------
    ws_cover = wb.active
    ws_cover.title = "1. 문서 개요 (Overview)"
    ws_cover.views.sheetView[0].showGridLines = True
    
    ws_cover.cell(row=2, column=2, value="Factory Operation AI Agent").font = Font(name=font_family, size=20, bold=True, color="1E3A8A")
    ws_cover.cell(row=3, column=2, value="정보구조도 (Information Architecture & Screen Hierarchy)").font = Font(name=font_family, size=14, bold=True, color="374151")
    ws_cover.cell(row=4, column=2, value="기반 산출물: UI Design Aligned.html (기준 문서 v0.13, 관제실 PC 1280×840 기준)").font = subtitle_font
    
    # Document Metadata Table
    meta_headers = ["항목", "내용"]
    meta_data = [
        ("시스템명", "Factory Operation AI Agent (공정 물류/설비 관제 AI 어시스턴트)"),
        ("설계 기준 해상도", "관제실 데스크톱 표준 1280 × 840 (3단 분할 레이아웃: LNB / 대화창 / 결과 패널)"),
        ("문서 버전", "v1.0 (2026-08-25 기준)"),
        ("추출 소스", "UI Design Aligned.html (화면 ID: S01 ~ S14 전체 14개 화면)"),
        ("주요 사용자", "보전 엔지니어 (Maintenance), 공정 물류 관리자 (Operations), 시스템 관리자 (Admin)"),
        ("보안 및 환경", "사내 온프레미스 폐쇄망, TLS 1.3 암호화, 역할 기반 권한 제어 (RBAC: User, Admin)")
    ]
    
    ws_cover.cell(row=6, column=2, value="■ 시스템 및 문서 정보").font = section_font
    
    for c_idx, h in enumerate(meta_headers, start=2):
        cell = ws_cover.cell(row=7, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
        
    for r_idx, (k, v) in enumerate(meta_data, start=8):
        c1 = ws_cover.cell(row=r_idx, column=2, value=k)
        c2 = ws_cover.cell(row=r_idx, column=3, value=v)
        c1.font = cell_font_bold
        c2.font = cell_font
        c1.alignment = align_center
        c2.alignment = align_left
        c1.fill = zebra_fill
        c1.border = cell_border
        c2.border = cell_border
        
    # IA Summary Statistics Table
    ws_cover.cell(row=16, column=2, value="■ 화면 분류 체계 및 구성 현황").font = section_font
    
    summary_headers = ["대분류 (1Depth)", "화면 수", "화면 ID 목록", "핵심 제공 기능"]
    summary_data = [
        ("A. 접속 (Authentication)", "1", "S01", "사번/비밀번호 로그인, 실패 5회 잠금, 세션 만료(30분) 재인증"),
        ("B. 질의 · 응답 (Query & Response)", "7", "S02 ~ S08", "메인 대화, 설비 가동률 분석(동적표/파레토), SOP 수순 안내, 근거 원문 뷰어, 세션 작업대, 멀티모달 PDF 파싱, 운영 산출물 생성, 예외 처리 3종"),
        ("C. 개인 공간 (Workspace & History)", "2", "S09 ~ S10", "프로젝트 워크스페이스 관리(전용 매뉴얼 인덱싱/검색 범위 토글), 대화 히스토리 검색/필터링/스냅샷(1년 보관)"),
        ("D. 관리자 (Admin Console)", "4", "S11 ~ S14", "관리자 콘솔(Vector DB 인덱스/답변불가 로그/임계값 설정/시스템 상태), 전사 지식 등록 모달, 계정·권한 관리(2단 그리드), 신규 계정 등록 모달")
    ]
    
    for c_idx, h in enumerate(summary_headers, start=2):
        cell = ws_cover.cell(row=17, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
        
    for r_idx, row in enumerate(summary_data, start=18):
        for c_idx, val in enumerate(row, start=2):
            cell = ws_cover.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font_bold if c_idx in [2, 3] else cell_font
            cell.alignment = align_center if c_idx in [2, 3, 4] else align_left
            cell.border = cell_border
            if r_idx % 2 == 0:
                cell.fill = zebra_fill
                
    ws_cover.column_dimensions['A'].width = 4
    ws_cover.column_dimensions['B'].width = 30
    ws_cover.column_dimensions['C'].width = 15
    ws_cover.column_dimensions['D'].width = 20
    ws_cover.column_dimensions['E'].width = 85
    
    # ---------------------------------------------------------
    # Sheet 2: 정보구조도 (IA - Information Architecture)
    # ---------------------------------------------------------
    ws_ia = wb.create_sheet(title="2. 정보구조도 (IA)")
    ws_ia.views.sheetView[0].showGridLines = True
    
    ws_ia.cell(row=2, column=2, value="Factory Operation AI Agent - 정보구조도 (Information Architecture)").font = title_font
    ws_ia.cell(row=3, column=2, value="UI Design Aligned.html에 정의된 화면 구조, 1~3Depth 계층, UI 컴포넌트, 인터랙션, 요구사항 추적 매핑").font = subtitle_font
    
    ia_headers = [
        "1Depth (대분류)", "2Depth (중분류)", "3Depth (소분류/기능명)", 
        "화면 ID", "화면명", "화면 유형", "접근 권한 (RBAC)", 
        "주요 UI 컴포넌트 / 영역", "주요 사용자 액션 및 인터랙션", 
        "연계 요구사항 / 표준", "비고 및 화면 특성"
    ]
    
    # Full IA Data
    ia_data = [
        # A. 접속
        ("A. 접속", "로그인", "사번 / 비밀번호 인증", "S01", "로그인 · 세션 재인증", "단독 화면", "All", 
         "• 시스템 타이틀/안내 영역\n• 온프레미스 폐쇄망/TLS 1.3 배지\n• 사번 입력 필드\n• 비밀번호 입력 필드 (마스킹/표시 토글)\n• 로그인 오류 메시지 & 실패 카운트 표시\n• 로그인 실행 버튼\n• 비밀번호 재설정 관리자 문의 안내",
         "• 사번/비밀번호 입력 후 [로그인] 클릭\n• 비밀번호 표시/숨김 토글\n• 5회 실패 시 계정 잠금 처리\n• 30분 무조작 세션 만료 시 재인증 안내 팝업 후 해당 화면으로 리다이렉트",
         "SR-001, SR-002, SR-003, SR-004", "SSO/소셜로그인 제외 (폐쇄망 사번 인증 전용)"),
        
        ("A. 접속", "세션 관리", "세션 만료 자동 재인증", "S01", "로그인 · 세션 재인증", "단독 화면", "All",
         "• 세션 만료 안내 배너 ('세션이 만료되어 다시 로그인이 필요합니다')\n• 이전 작업 컨텍스트 보존 알림",
         "• 무조작 30분 경과 시 자동 세션 만료 처리\n• 재로그인 성공 시 이전 작업 화면/대화 세션 복원",
         "SR-004", "보안 정책에 따른 30분 타임아웃"),

        # B. 질의 · 응답
        ("B. 질의 · 응답", "메인 대화", "설비 가동률 분석 질의", "S02", "메인 채팅 · 설비 가동률 분석 응답", "3단 레이아웃 (기본)", "User, Admin",
         "• LNB: 메뉴(대화/워크스페이스/히스토리/관리자), 최근 질의 퀵링크, 프로필\n• 중앙 대화창: 질의/답변 버블, SLM 처리과정(Clas-Plan-Proc) 아코디언, 퀵 프롬프트 칩\n• 우측 결과 패널: 동적 표(호기별 가동률/비가동), 파레토 차트(Top3 원인), AI 개선 권고 카드, 출처 표시",
         "• 자연어 질의 입력 및 전송\n• 처리과정(Multi-SLM 파이프라인) 펼치기/접기\n• 퀵 프롬프트 칩 클릭 (예: '3호기만 다시 분석', '지난주와 비교')\n• 우측 동적 표/차트 조회 및 탭 전환(동적 표/요약 카드/출처)\n• AI 개선 권고 카드 내 [PM 체크리스트 초안 생성] 버튼 클릭",
         "FR-001, FR-002, FR-010, Relay API", "3단 분할 레이아웃(184px:452px:나머지), 인라인 표출 원칙 대응"),

        ("B. 질의 · 응답", "메인 대화", "SOP 단계별 가이드 & 원문 조회", "S03", "SOP 단계별 응답 · 근거 원문 뷰어", "3단 레이아웃 (뷰어 전환)", "User, Admin",
         "• LNB: 최근 질의 퀵링크\n• 중앙 대화창: 안전 주의사항 배너(LOTO 필수), 단계별 조치 수순 카드(1, 2단계), 인용 출처 링크 배너\n• 우측 결과 패널 (원문 뷰어): PDF 뷰어 헤더(파일명/유사도/원문 다운로드), 문서 본문 렌더링, 인용 청크 강조 하이라이팅, 도면 캡처 영역, 페이지 이동(◀/▶)",
         "• SOP 질의 입력 (예: 'ERR-104 조치 SOP')\n• 답변 내 단계별 수순 확인\n• '우측 표시 중' 출처 클릭 시 패널에 해당 페이지/청크 즉시 렌더링\n• 뷰어 내 이전/다음 페이지 이동\n• [원문 다운로드] 클릭",
         "FR-021, FN-SOP-01~03", "패널이 원문 뷰어로 동적 전환, 유사도(0.86) 및 인용 청크 하이라이팅 표출"),

        ("B. 질의 · 응답", "메인 대화", "텍스트 단독 응답 (패널 접힘)", "S04", "텍스트 응답 시 패널 접힘 (대안)", "2단 레이아웃 (중앙 확장)", "User, Admin",
         "• LNB\n• 대화창 상단: 대화 제목, [결과 패널 열기 ◧] 토글 버튼\n• 중앙 대화창 (최대 720px 확장): 텍스트 답변, 정비기준 강조, RAG 인용 출처 카드, 연관 퀵 프롬프트 칩\n• 하단 입력부: 파일첨부, 음성입력, 전송",
         "• 단답형/텍스트 중심 질의 입력\n• 패널 자동 접힘으로 가독 행폭(720px) 확보\n• 필요 시 상단 [결과 패널 열기] 버튼 클릭으로 수동 확장\n• RAG 인용 카드 [원문 열기 ↗] 클릭 시 뷰어 패널 재오픈",
         "FR-021, UI/UX 최적화", "정형 데이터 부재 시 가독성 극대화를 위한 적응형 레이아웃 대안"),

        ("B. 질의 · 응답", "세션 작업대", "결과물 누적 및 교차 참조", "S05", "세션 작업대 · 직전 결과 누적", "3단 레이아웃 (작업대)", "User, Admin",
         "• 중앙 대화창: 연속 질의 흐름 (원인 분석 → 산출물 작성 요청), 생성 산출물 링크 카드\n• 우측 결과 패널 (세션 작업대): 대화 결과물 스택(3건), PM 점검표 초안 인라인 그리드, 유사 장애 조치 이력 카드, 직전 가동률 집계 축약 카드([펼치기 ▾])",
         "• 이전 분석 결과를 참조하여 연속 프롬프트 실행 (예: '방금 표에서 3호기 원인 찾아줘')\n• 패널 내 생성된 산출물([PM_RGV_주간_초안]) 검토 및 인라인 편집\n• [표준 문서로 등록] 클릭\n• 직전 결과 카드 펼침/접힘 토글",
         "FR-015, FN-ADV-01", "세션 내 생성된 데이터/산출물이 우측 패널에 카드 형태로 누적 보존"),

        ("B. 질의 · 응답", "멀티모달 질의", "PDF 업로드 및 인메모리 파싱", "S06", "멀티모달 PDF 업로드 · 파싱", "3단 레이아웃 (파싱/임베딩)", "User, Admin",
         "• 중앙 대화창: 업로드된 첨부파일 칩(파일명/용량/파싱상태/삭제), 임시 RAG 처리과정(OCR/인메모리 임베딩), 3단계 수순 답변\n• 하단 입력부: 파일 파싱 진행률 프로그레스바(%), 지원 포맷 안내(50MB 이하)\n• 우측 결과 패널: 업로드 문서 원문 뷰어, 배선도/도면 캡처 영역, [영구 보관(워크스페이스)] 버튼",
         "• [+ 파일] 클릭 또는 드래그앤드롭으로 PDF/도면 업로드\n• 실시간 파싱/임베딩 진행률 확인\n• 첨부 문서 기반 질의 수행 (예: '에러코드 E-203 조치법 요약')\n• 세션 종료 전 [영구 보관(워크스페이스)] 클릭으로 프로젝트에 등록",
         "FN-SOP-04, FR-006, DR-004", "세션 종료 시 임베딩 임시 폐기 원칙 (영구 보관 선택 가능)"),

        ("B. 질의 · 응답", "운영 산출물", "PM / 시운전 / BCP 산출물 생성", "S07", "운영 산출물 생성 (PM/시운전/BCP)", "3단 레이아웃 (에디터)", "User, Admin",
         "• 중앙 대화창: 복수 산출물 생성 처리과정, 초안 링크 카드 2종, 참조 근거 표시, 변환 퀵 칩\n• 우측 결과 패널: 탭 네비(PM 체크리스트 / 시운전 5단계), 주간 PM 점검표 편집 가능 그리드, 행 추가/항목 편집 링크, 표준 마스터 데이터 제안 카드, [표준 문서로 등록] 버튼",
         "• 산출물 생성 프롬프트 실행 (예: '신규 도입 RGV 5호기 PM 체크리스트 작성')\n• 패널 탭 전환으로 여러 산출물(PM표/시운전가이드) 교차 검토\n• 그리드 내 판정 기준값 인라인 수정 및 행 추가\n• [표준 문서로 등록] 클릭으로 마스터 지식 DB 반영 요청",
         "FN-ADV-01~03, FR-015~017", "검토자 승인 워크플로우 연계, 표준 마스터 매핑 자동 제안"),

        ("B. 질의 · 응답", "예외 처리", "필수 인자 누락 보완", "S08-1", "예외 흐름 1: 필수 인자 누락", "대화 컬럼", "User, Admin",
         "• 질의 부족 감지 안내 메시지\n• 설비 선택 칩 ([1] AGV 3호기 ERR-104, [2] S/C 1호기 AL-231)\n• [직접 입력하기] 선택 칩\n• 하단 입력 힌트",
         "• 불명확한 질의 입력 시 시스템이 현재 알람 설비 칩을 추천\n• 칩 클릭 시 즉시 인자 보완 후 RAG 탐색으로 자동 전이\n• 직접 입력을 통해 텍스트로 보완 입력",
         "FR-020, Clas-SLM", "에러 팝업을 띄우지 않고 대화형 칩 선택으로 매끄러운 흐름 유도"),

        ("B. 질의 · 응답", "예외 처리", "RAG 근거 부족 (유사도 미달)", "S08-2", "예외 흐름 2: RAG 근거 부족", "대화 컬럼", "User, Admin",
         "• '답변할 수 없습니다' 안내 카드 (붉은색/경고 테마)\n• 검색 결과 메타데이터 (최고 유사도 0.41 / 기준 0.70 미달)\n• 후속 조치 액션 칩 ([매뉴얼 PDF 첨부해 다시 질의], [매뉴얼 원본 확인·담당 부서 문의])",
         "• 근거 부족 시 환각(Hallucination) 방지를 위해 즉시 답변 거부\n• 유사도 수치 및 검색 범위 확인\n• 권고 칩 클릭으로 신규 매뉴얼 업로드 또는 수동 절차로 전환",
         "FR-022, 2.2.2 Hallucination 방지", "유사도 임계값(0.70) 미달 시 임의 생성 엄격 차단"),

        ("B. 질의 · 응답", "예외 처리", "중계 서버 통신 장애 (타임아웃)", "S08-3", "예외 흐름 3: 중계 서버 통신 장애", "대화 컬럼", "User, Admin",
         "• 처리과정 내 Relay API 타임아웃(3.0s) 에러 표시\n• '중계 서버 통신 실패' 안내 카드\n• 하단 상태바 (중계 서버 연결 끊김 인디케이터, 재연결 시도 중)\n• 액션 칩 ([다시 시도], [지식 질의로 계속 (RAG는 정상)])",
         "• 중계 서버(SCADA/WMS) 3초 타임아웃 발생 시 현장 관제실 화면 확인 안내\n• [다시 시도] 클릭\n• [지식 질의로 계속] 선택 시 설비 데이터 연계 없이 내부 Vector DB RAG 질의 지속",
         "FR-020, Relay API Fail-Safe", "로컬 큐 버퍼링 지원, Vector DB 지식 질의는 독립적으로 정상 작동"),

        # C. 개인 공간
        ("C. 개인 공간", "워크스페이스", "프로젝트 워크스페이스 관리", "S09", "프로젝트 워크스페이스 관리", "2단 그리드/목록", "User, Admin",
         "• LNB\n• 좌측 패널: 내 프로젝트 목록(3건), [＋ 새 프로젝트], 프로젝트별 문서/청크 수, '검색 범위로 사용' 토글 스위치(ON/OFF), 프로젝트 관리 칩(문서추가/이름변경/공유), 파일 드래그앤드롭 업로드존\n• 우측 패널: 선택된 프로젝트 문서 인덱스 테이블(문서명/등록일/페이지/인덱스상태), 검색 범위 연동 스위치, [대화 열기] 버튼",
         "• [＋ 새 프로젝트] 클릭으로 전용 작업공간 생성\n• 매뉴얼/도면 파일 드래그앤드롭 업로드 (50MB 이하)\n• 특정 프로젝트의 '검색 범위로 사용' 스위치 토글 ON/OFF (맞춤 검색 지정)\n• 프로젝트 내 문서 파싱/인덱싱 상태 확인\n• [이 프로젝트로 새 대화 시작] 클릭 시 검색 범위 칩이 지정된 대화방으로 진입",
         "FN-PWS-01~02, FR-018~019", "프로젝트별 전용 매뉴얼 인덱싱 및 맞춤형 검색 스코프 격리"),

        ("C. 개인 공간", "히스토리", "대화 이력 조회 및 검색", "S10", "히스토리 조회 · 검색", "2단 분할/미리보기", "User, Admin",
         "• LNB\n• 좌측 패널: 키워드 검색창, 유형별 필터 칩(전체/SOP/설비 분석/재고/산출물/답변 불가), 일자별(오늘/어제/이전) 대화 목록, 대화 요약/시간/태그\n• 우측 패널 (스냅샷 미리보기): 선택 대화 제목, [이어서 질의] 버튼, 질의 원문 & 실행 메타(시간/SLM/Relay API), 응답 스냅샷(표/텍스트), 참조 근거 로그",
         "• 키워드(설비명, 알람코드 등) 검색 및 카테고리 필터 선택\n• 대화 항목 클릭 시 우측에 당시 응답 스냅샷 및 근거 로그 실시간 프리뷰\n• [이어서 질의] 클릭 시 해당 세션 컨텍스트를 유지한 채 메인 대화창으로 전환",
         "FR-001, DR-003, DR-006", "대화 이력 최소 1년 보관 준수, 응답 당시 스냅샷 및 근거 로그 완벽 추적"),

        # D. 관리자
        ("D. 관리자", "관리자 콘솔", "시스템 및 지식 인덱스 모니터링", "S11", "관리자 콘솔", "관리자 대시보드", "Admin",
         "• LNB (Admin 배지 표시)\n• Vector DB 지식 인덱스 현황 테이블(지식군/개정/청크/상태), [＋ 문서 등록] 버튼\n• 답변 불가 / 근거 부족 로그 카드 (최근 7일 Top 실패 질의 집계)\n• 신뢰도 기준값(유사도) 슬라이더 조절 UI (SOP 0.70, 트러블슈팅 0.65)\n• 계정/권한 요약 카드 (User/Admin/잠금수), [계정 목록 관리] 버튼\n• 시스템 상태 모니터링 카드 (SLM/중계서버 지연/Vector DB/최종백업)",
         "• 전사 지식군별 인덱스 현황 및 노후화 상태 점검\n• 답변 불가 질의 로그 분석을 통한 지식 등록 필요성 파악 (예: 팔레타이저 P-9)\n• 질의 유형별 RAG 유사도 임계값 슬라이더로 실시간 조정\n• [＋ 문서 등록] 클릭으로 S12 모달 호출\n• [계정 목록 관리] 클릭으로 S13 화면 이동",
         "SR-001, OR-004, 5.7 지식 노후화 방지", "Admin 전용 콘솔, 지식 상태 및 시스템 가동 지표 종합 관제"),

        ("D. 관리자", "전사 지식 관리", "전사 표준 지식 문서 등록", "S12", "전사 지식 문서 등록 (모달)", "모달 (Modal)", "Admin",
         "• 모달 헤더 ([전사 지식 문서 등록], [×] 닫기)\n• 파일 드래그앤드롭 업로드존 (지원 포맷/용량 50MB 안내)\n• 업로드 파일 목록 & 실시간 진행률 바, 미지원 파일 경고 표시\n• 지식군 드롭다운 선택 필드 (*필수)\n• 개정 버전 입력 필드 (*필수)\n• [기존 버전 대체 및 최신 지식 우선 검색] 체크박스\n• 등록 후 파이프라인 안내 (텍스트/도표 파싱 → 청크 임베딩 → 색인 → 백업)\n• [취소], [등록 및 색인] 버튼",
         "• 표준 매뉴얼/SOP 파일 드래그앤드롭 업로드\n• 지식군 선택 및 개정 버전(v2.0) 입력\n• 이전 버전 대체 여부 체크 (지식 노후화 방지)\n• [등록 및 색인] 클릭 시 파싱/임베딩 백그라운드 작업 시작 및 Vector DB 색인",
         "OR-004, DR-001, DR-004", "S11 관리자 콘솔 위 레이어 모달, 다중 포맷(PDF/Office/이미지 등) 지원"),

        ("D. 관리자", "계정 / 권한 관리", "사용자 계정 및 보안 정책 관리", "S13", "관리자 · 계정 및 권한 관리", "2단 그리드/상세", "Admin",
         "• LNB\n• 좌측 패널: 사용자 검색창(이름/사번/소속), 권한/상태 필터 칩(전체/User/Admin/잠금/비활성), 계정 목록 테이블(146명), [＋ 계정 등록] 버튼\n• 우측 패널: 선택 계정 상세 정보, RBAC 권한 토글/선택(User/Admin), 계정 상태(마지막 접속/실패 횟수/비밀번호 변경일), [비밀번호 초기화] / [계정 비활성] 버튼, 잠금 계정 알림 카드 & [잠금 해제] 버튼, 공통 보안 정책 요약(30분 세션/90일 변경/5회 잠금)",
         "• 사번/이름 검색 및 상태 필터링\n• 사용자 클릭 시 우측 상세 패널에 세부 정보 로드\n• 사용자 권한(RBAC) 변경 (User ↔ Admin)\n• 실패 횟수 초과 계정 [잠금 해제] 실행\n• [비밀번호 초기화] 실행\n• [＋ 계정 등록] 클릭 시 S14 모달 호출",
         "SR-001, SR-003, SR-004, DR-002", "계정 관리 및 RBAC 권한 통제, 잠금 해제 기능 제공"),

        ("D. 관리자", "계정 등록", "신규 사용자 계정 생성", "S14", "계정 등록 (모달)", "모달 (Modal)", "Admin",
         "• 모달 헤더 ([계정 등록], [×] 닫기)\n• 이름 입력 필드 (*필수)\n• 사번 (로그인 ID) 입력 필드 (*필수)\n• 소속 부서 드롭다운 (*필수)\n• 권한(RBAC) 선택 라디오 카드 (User / Admin)\n• 초기 비밀번호 입력 필드 (*필수)\n• [최초 로그인 시 비밀번호 변경 요구] 체크박스\n• 비밀번호 정책 안내 (10자 이상, 영문/숫자/특수문자 조합, 90일 주기)\n• [취소], [등록] 버튼",
         "• 신규 입사자/운영자 정보 입력\n• 소속 및 권한(User/Admin) 지정\n• 초기 비밀번호 설정 및 변경 강제 체크\n• [등록] 클릭 시 유효성 검증 후 DB 저장 및 S13 목록에 즉시 반영",
         "SR-001, SR-003, DR-002", "S13 계정 관리 위 레이어 모달, 필수값 및 비밀번호 복잡도 검증")
    ]
    
    for c_idx, h in enumerate(ia_headers, start=2):
        cell = ws_ia.cell(row=5, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
        
    for r_idx, row in enumerate(ia_data, start=6):
        cat = row[0]
        fill_color = zebra_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        if "A. 접속" in cat:
            cat_fill = category_a_fill
        elif "B. 질의" in cat:
            cat_fill = category_b_fill
        elif "C. 개인" in cat:
            cat_fill = category_c_fill
        elif "D. 관리자" in cat:
            cat_fill = category_d_fill
        else:
            cat_fill = fill_color
            
        for c_idx, val in enumerate(row, start=2):
            cell = ws_ia.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.border = cell_border
            
            # Alignments & Styles
            if c_idx in [2, 3]: # 1Depth, 2Depth
                cell.font = cell_font_bold
                cell.alignment = align_center
                cell.fill = cat_fill
            elif c_idx == 4: # 3Depth
                cell.font = cell_font_bold
                cell.alignment = align_left
                cell.fill = fill_color
            elif c_idx == 5: # Screen ID
                cell.font = Font(name="Consolas", size=9.5, bold=True, color="1E3A8A")
                cell.alignment = align_center
                cell.fill = fill_color
            elif c_idx in [6, 7, 8]: # 화면명, 유형, 권한
                cell.font = cell_font_bold if c_idx == 6 else cell_font
                cell.alignment = align_center if c_idx in [7, 8] else align_left
                cell.fill = fill_color
            elif c_idx in [9, 10]: # UI 컴포넌트, 액션
                cell.alignment = align_left
                cell.fill = fill_color
            elif c_idx == 11: # 요구사항 매핑
                cell.font = cell_font_mono
                cell.alignment = align_center
                cell.fill = fill_color
            else: # 비고
                cell.alignment = align_left
                cell.fill = fill_color

    ws_ia.column_dimensions['A'].width = 3
    ws_ia.column_dimensions['B'].width = 16  # 1Depth
    ws_ia.column_dimensions['C'].width = 16  # 2Depth
    ws_ia.column_dimensions['D'].width = 24  # 3Depth
    ws_ia.column_dimensions['E'].width = 12  # Screen ID
    ws_ia.column_dimensions['F'].width = 28  # Screen Name
    ws_ia.column_dimensions['G'].width = 20  # Screen Type
    ws_ia.column_dimensions['H'].width = 14  # RBAC
    ws_ia.column_dimensions['I'].width = 45  # UI Components
    ws_ia.column_dimensions['J'].width = 45  # Actions & Interactions
    ws_ia.column_dimensions['K'].width = 24  # Traceability
    ws_ia.column_dimensions['L'].width = 35  # Notes

    # ---------------------------------------------------------
    # Sheet 3: 화면 목록 정의서 (Screen List Spec)
    # ---------------------------------------------------------
    ws_screens = wb.create_sheet(title="3. 화면 목록 정의서 (Screens)")
    ws_screens.views.sheetView[0].showGridLines = True
    
    ws_screens.cell(row=2, column=2, value="Factory Operation AI Agent - 화면 목록 정의서 (Screen List)").font = title_font
    ws_screens.cell(row=3, column=2, value="14개 전체 화면(S01~S14)의 레이아웃 구조, UI 목적, 상세 컴포넌트 및 설계 근거").font = subtitle_font
    
    screen_headers = [
        "화면 ID", "화면 명칭", "대분류", "레이아웃 방식", 
        "화면 UI 목적 및 주요 기능", "주요 구성 요소 (좌측/중앙/우측)", 
        "설계 의도 및 가이드라인 (Design Rationale)", "기준 URS/요구사항"
    ]
    
    screens_data = [
        ("S01", "로그인 · 세션 재인증", "A. 접속", "단독 중앙 카드 (1000px flex)",
         "사용자 사번/비밀번호 인증 및 30분 무조작 세션 만료 시 재인증 처리",
         "좌측: 시스템 브랜딩, 폐쇄망/TLS 1.3 상태 배지\n우측: 사번/PW 입력, 실패 카운터, 관리자 문의 안내",
         "SSO 및 비밀번호 자가 재설정은 사내 보안 규정에 따라 배제. 5회 실패 잠금 및 온프레미스 폐쇄망 보안 준수.",
         "SR-001, SR-002, SR-003, SR-004"),
        
        ("S02", "메인 채팅 · 설비 가동률 분석 응답", "B. 질의·응답", "3단 고정 (184px : 452px : 가변)",
         "자연어 질의를 통한 설비 가동률/비가동 분석 및 동적 표/차트/개선권고 표출",
         "좌: LNB(메뉴, 최근질의, 프로필)\n중: 대화창(Multi-SLM 처리과정, 답변, 퀵칩)\n우: 동적 표, 파레토 차트, AI 개선권고 카드, 출처",
         "넓은 화면(1280px)을 활용하여 우측에 표를 고정(Pin)해 두고 대화를 지속 가능. 관리자 분석 리포트 작성에 최적화.",
         "FR-001, FR-002, FR-010, Relay API"),

        ("S03", "SOP 단계별 응답 · 근거 원문 뷰어", "B. 질의·응답", "3단 고정 (패널 원문 뷰어 전환)",
         "장애/알람 발생 시 단계별 SOP 수순 안내 및 우측 패널 근거 원문 PDF 즉시 렌더링",
         "좌: LNB\n중: 안전 수칙(LOTO), 단계별 조치 카드, 인용 출처\n우: PDF 원문 뷰어, 청크 하이라이팅, 도면 캡처, 페이지 네비",
         "그라운딩(FR-021)을 눈으로 직접 확인시켜 현장 엔지니어의 신뢰도 극대화. 유사도(0.86) 및 인용 페이지 직결.",
         "FR-021, FN-SOP-01~03"),

        ("S04", "텍스트 응답 시 패널 접힘 (대안)", "B. 질의·응답", "2단 가변 (중앙 720px 확장)",
         "정형 데이터/문서 뷰어가 필요 없는 단순 텍스트 질의 시 대화 가독성 확보",
         "좌: LNB\n중: 720px 최적 행폭 대화창, RAG 인용 카드, [결과 패널 열기 ◧] 버튼\n우: 접힘 상태",
         "응답 성격에 따라 레이아웃이 자동 적응. 텍스트 응답 시 시선 분산을 막고 안정된 행폭 유지.",
         "FR-021, UI/UX 반응형"),

        ("S05", "세션 작업대 · 직전 결과 누적", "B. 질의·응답", "3단 고정 (우측 스택형 누적)",
         "연속 질의를 통해 생성된 복수의 분석 결과 및 산출물을 우측 패널에 카드 형태로 누적",
         "좌: LNB\n중: 연속 질의 대화 버블\n우: 세션 결과물 스택(PM 점검표 인라인 편집, 유사장애 이력, 가동률 집계 축약)",
         "질의가 이어져도 직전 분석 표와 산출물이 우측에 유지되어 대화 중 교차 참조가 용이함. 관리자 작업 흐름에 최적.",
         "FR-015, FN-ADV-01"),

        ("S06", "멀티모달 PDF 업로드 · 파싱", "B. 질의·응답", "3단 고정 (인메모리 임베딩 연계)",
         "사용자 매뉴얼/도면 PDF 업로드, 실시간 OCR/파싱 및 세션 인메모리 RAG 탐색",
         "좌: LNB\n중: 첨부 파일 카드, 임베딩 진행과정, 3단계 조치 수순\n우: 업로드 도면/배선도 원문 뷰어, [영구 보관] 버튼",
         "세션 종료 시 임베딩이 자동 폐기되는 임시 분석 기능. 필요 시 프로젝트 워크스페이스에 영구 등록 가능.",
         "FN-SOP-04, FR-006, DR-004"),

        ("S07", "운영 산출물 생성 (PM/시운전/BCP)", "B. 질의·응답", "3단 고정 (우측 인라인 에디터)",
         "설비 도입 및 운영에 필요한 PM 체크리스트, 시운전 5단계 절차서, BCP 문서 자동 생성 및 편집",
         "좌: LNB\n중: 2종 산출물 생성 버블, 참조 근거\n우: 탭 전환 뷰(PM 점검표/시운전), 그리드 인라인 편집기, 표준 등록 버튼",
         "AI가 초안을 생성하고 사용자가 우측 패널에서 직접 기준값을 수정한 후 표준 마스터로 즉시 확정 등록.",
         "FN-ADV-01~03, FR-015~017"),

        ("S08", "예외 흐름 3종 처리", "B. 질의·응답", "대화 컬럼 3단 비교 (화면 발췌)",
         "① 필수 인자 누락 시 대화형 칩 보완\n② RAG 유사도 미달(<0.70) 시 답변 거부\n③ 중계 서버 통신 실패(3s) 시 안내 및 RAG 지속",
         "컬럼 1: 필수 인자 추천 칩\n컬럼 2: 답변 거부 및 유사도(0.41) 표시, 권고 칩\n컬럼 3: Relay API 타임아웃 피드백, 연결 상태바",
         "에러 발생 시에도 시스템이 중단되지 않고 대화로 복구하거나 명확한 신뢰도 기준을 제시하여 안전 운영 보장.",
         "FR-020, FR-022, 2.2.2"),

        ("S09", "프로젝트 워크스페이스 관리", "C. 개인공간", "2단 분할 (목록 / 문서 인덱스)",
         "개인 및 반별 전용 프로젝트 생성, 맞춤형 매뉴얼 등록/인덱싱 및 검색 범위(Scope) 설정",
         "좌: 내 프로젝트 목록, 검색 범위 토글 스위치, 업로드존\n우: 선택 프로젝트 문서 인덱스 테이블, [대화 열기] 버튼",
         "특정 호기/라인 전용 매뉴얼만 격리하여 맞춤 RAG 탐색을 수행할 수 있도록 검색 스코프 제어 지원.",
         "FN-PWS-01~02, FR-018~019"),

        ("S10", "히스토리 조회 · 검색", "C. 개인공간", "2단 분할 (목록 / 스냅샷 프리뷰)",
         "과거 대화 이력의 다차원 키워드/카테고리 검색 및 당시 응답 스냅샷/근거 로그 완벽 복원",
         "좌: 검색창, 카테고리 필터 칩, 일자별 대화 목록\n우: 선택 대화 스냅샷(질의/응답 표/참조 근거/Relay API 로그), [이어서 질의]",
         "1년치 대화 이력 보관(DR-003) 규정을 준수하며, 감사 및 재현이 가능하도록 당시 상태 스냅샷 제공.",
         "FR-001, DR-003, DR-006"),

        ("S11", "관리자 콘솔", "D. 관리자", "관리자 대시보드 그리드",
         "Vector DB 지식 인덱스 상태 점검, 답변 불가 로그 분석, RAG 유사도 임계값 조정 및 시스템 상태 모니터링",
         "좌측 영역: Vector DB 지식 현황 테이블, 답변 불가 로그(7일)\n우측 영역: 유사도 임계값 슬라이더, 계정 통계, 시스템 헬스체크",
         "지식 노후화를 방지(OR-004, 5.7)하고 임계값을 실시간 조정할 수 있는 중앙 관제 환경 제공 (Admin 전용).",
         "SR-001, OR-004, 5.7"),

        ("S12", "전사 지식 문서 등록 (모달)", "D. 관리자", "모달 (S11 오버레이 레이어)",
         "신규 매뉴얼/SOP/도면 파일 업로드, 지식군/개정 버전 지정 및 최신 지식 우선 색인 파이프라인 가동",
         "중앙 모달: 드래그앤드롭 업로드존, 진행률 바, 지식군 선택, 개정버전 입력, [기존 버전 대체] 체크, [등록 및 색인]",
         "최신 개정 버전 등록 시 구버전을 자동 대체하여 AI가 항상 최신 표준 지식을 참조하도록 보장.",
         "OR-004, DR-001, DR-004"),

        ("S13", "관리자 · 계정 및 권한 관리", "D. 관리자", "2단 분할 (계정 목록 / 상세 패널)",
         "전체 사용자(146명) 계정 조회, RBAC 권한(User/Admin) 변경, 5회 실패 잠금 해제, 비밀번호 초기화",
         "좌: 계정 검색 및 상태 필터 칩, 사용자 목록 테이블\n우: 선택 사용자 상세, RBAC 토글, 계정 상태, [잠금 해제], 보안 정책 요약",
         "조직도 연동 등 불필요한 기능은 배제하고 폐쇄망 환경에 필수적인 사번 계정 관리 및 잠금 해제에 집중.",
         "SR-001, SR-003, SR-004, DR-002"),

        ("S14", "계정 등록 (모달)", "D. 관리자", "모달 (S13 오버레이 레이어)",
         "신규 사용자 사번, 이름, 소속, RBAC 권한, 초기 비밀번호 설정 및 최초 로그인 시 변경 강제",
         "중앙 모달: 이름/사번/소속/권한 라디오/초기PW 입력 필드, [최초 로그인 시 변경 요구] 체크, 비밀번호 정책 안내",
         "사내 보안 정책(10자 이상 복잡도, 90일 주기)을 준수하는 표준 계정 생성 모달.",
         "SR-001, SR-003, DR-002")
    ]
    
    for c_idx, h in enumerate(screen_headers, start=2):
        cell = ws_screens.cell(row=5, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
        
    for r_idx, row in enumerate(screens_data, start=6):
        fill_color = zebra_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row, start=2):
            cell = ws_screens.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.border = cell_border
            cell.fill = fill_color
            
            if c_idx == 2: # Screen ID
                cell.font = Font(name="Consolas", size=10, bold=True, color="1E3A8A")
                cell.alignment = align_center
            elif c_idx in [3, 4]: # Screen Name, Category
                cell.font = cell_font_bold
                cell.alignment = align_center if c_idx == 4 else align_left
            elif c_idx == 5: # Layout
                cell.alignment = align_center
            elif c_idx in [6, 7, 8]: # Purpose, Components, Rationale
                cell.alignment = align_left
            elif c_idx == 9: # Requirements
                cell.font = cell_font_mono
                cell.alignment = align_center
                
    ws_screens.column_dimensions['A'].width = 3
    ws_screens.column_dimensions['B'].width = 12  # Screen ID
    ws_screens.column_dimensions['C'].width = 28  # Screen Name
    ws_screens.column_dimensions['D'].width = 15  # Category
    ws_screens.column_dimensions['E'].width = 25  # Layout
    ws_screens.column_dimensions['F'].width = 42  # Purpose
    ws_screens.column_dimensions['G'].width = 45  # Components
    ws_screens.column_dimensions['H'].width = 50  # Rationale
    ws_screens.column_dimensions['I'].width = 28  # Traceability

    # ---------------------------------------------------------
    # Sheet 4: UI 컴포넌트 및 인터랙션 명세 (UI Components Spec)
    # ---------------------------------------------------------
    ws_comp = wb.create_sheet(title="4. UI 컴포넌트 명세 (Components)")
    ws_comp.views.sheetView[0].showGridLines = True
    
    ws_comp.cell(row=2, column=2, value="Factory Operation AI Agent - 주요 UI 컴포넌트 및 인터랙션 명세").font = title_font
    ws_comp.cell(row=3, column=2, value="와이어프레임에서 공통 및 핵심 화면에 사용된 주요 UI 컴포넌트의 기능 및 동작 규칙 정의").font = subtitle_font
    
    comp_headers = [
        "컴포넌트 명", "적용 화면", "컴포넌트 유형", 
        "세부 UI 구성 요소", "사용자 인터랙션 및 상태 변화", "연계 기능 및 시스템 규칙"
    ]
    
    comp_data = [
        ("LNB (공통 좌측 내비게이션)", "S02 ~ S07, S09, S10, S11, S13", "공통 레이아웃 (184px)",
         "• 시스템 로고 및 명칭\n• 메인 메뉴 (대화, 워크스페이스, 히스토리, 관리자)\n• 최근 질의 퀵링크 목록 (최대 3~5건)\n• 사용자 프로필 아바타 및 역할 배지",
         "• 메뉴 클릭 시 해당 화면으로 즉시 라우팅\n• 현재 활성 메뉴 배경 하이라이팅 (#E8E5DF)\n• 최근 질의 클릭 시 해당 대화방 로드\n• Admin 권한 보유 시에만 '관리자' 메뉴 노출",
         "RBAC 기반 메뉴 노출 제어, 대화 세션 컨텍스트 유지"),

        ("Multi-SLM 처리과정 아코디언", "S02, S03, S04, S06, S07, S08", "대화창 내 정보 표출",
         "• 처리과정 헤더 (펼침/접힘 토글 아이콘, 소요 시간)\n• 파이프라인 단계 (① Clas-SLM 의도분류 → ② Plan-SLM 인자추출 → ③ Relay/RAG 데이터조회 → ④ Proc-SLM 서식화)",
         "• '▾ 처리 과정' 클릭 시 세부 SLM 추론 및 통신 로그 펼침/접힘\n• 에러/지연 발생 시 해당 단계 붉은색/주황색 하이라이팅",
         "투명한 AI 추론 과정(Explainable AI) 제공으로 신뢰도 확보"),

        ("퀵 프롬프트 칩 (Quick Chips)", "S02, S03, S04, S07, S08", "대화창 액션 컨트롤",
         "• 둥근 알약 형태 버튼 (예: '3호기만 다시 분석', '지난주와 비교', 'PM 체크리스트 초안 생성', '다시 시도')",
         "• 칩 클릭 시 해당 텍스트가 자동으로 입력창에 설정되거나 즉시 프롬프트로 전송됨\n• 호버 시 보더 색상 진화",
         "추천 후속 질의를 통해 작업 효율 극대화"),

        ("동적 표 및 파레토 차트 (Dynamic Table & Chart)", "S02, S05, S10", "결과 패널 위젯",
         "• 동적 데이터 그리드 (호기, 운전시간, 비가동시간, 가동률)\n• 비가동 원인 파레토 바 차트 (원인별 % 비중 및 프로그레스바)\n• AI 개선 권고 하이라이트 박스",
         "• 가동률 저하 설비 행 강조 배경색 (#FAF7F0)\n• 파레토 차트 마우스 호버 시 세부 수치 툴팁\n• 개선 권고 박스 내 액션 버튼 클릭 시 산출물 생성 자동 연계",
         "중계 서버(Relay API) SCADA 데이터 실시간 렌더링"),

        ("근거 원문 PDF 뷰어 (Grounding Viewer)", "S03, S06", "결과 패널 뷰어",
         "• 뷰어 상단 메타 (파일명, 페이지/전체페이지, 유사도 스코어, [원문 다운로드])\n• 본문 텍스트 렌더링 영역\n• 인용 청크 하이라이팅 박스\n• 도면/배선도 캡처 뷰어 영역\n• 하단 페이지 네비게이션 (◀ p.33 / p.35 ▶)",
         "• 대화창의 인용 출처 링크 클릭 시 해당 페이지로 즉시 이동\n• 이전/다음 페이지 이동 버튼 클릭\n• 도면 영역 클릭 시 원본 해상도 확대 팝업",
         "FR-021 그라운딩 검증, 인용 청크 일치율 표출"),

        ("산출물 인라인 에디터 (Artifact Editor)", "S05, S07", "결과 패널 에디터",
         "• 탭 네비게이션 (PM 점검표 / 시운전 5단계)\n• 체크리스트 점검표 그리드 (No, 점검부위, 점검방법, 판정기준)\n• [＋ 행 추가 / 항목 편집] 버튼\n• 표준 마스터 매핑 제안 박스\n• [표준 문서로 등록] 액션 버튼",
         "• 테이블 셀 클릭 후 판정 기준값 직접 수정\n• 행 추가 버튼 클릭으로 점검 항목 삽입\n• [표준 문서로 등록] 클릭 시 검토자 승인 팝업 호출 후 Vector DB 반영",
         "FN-ADV-01~03 마스터 데이터 제안 및 표준 등록 연계"),

        ("파일 업로드존 & 파싱 인디케이터", "S06, S09, S12", "입력 및 모달 컴포넌트",
         "• 드래그앤드롭 영역 (점선 보더, 지원 포맷/50MB 안내)\n• 업로드 파일 카드 (파일명, 용량, 프로그레스바, 삭제[×] 버튼)\n• 파싱 및 임베딩 상태 텍스트 (OCR 진행률, 청크 수)",
         "• 파일 드래그앤드롭 또는 클릭하여 탐색기 파일 선택\n• 실시간 업로드 및 OCR/파싱 진행률(%) 바 표시\n• 미지원 확장자 또는 용량 초과 시 붉은색 경고 박스 노출",
         "DR-004 다중 포맷 지원 (PDF/Office/이미지 등), 50MB 제한"),

        ("프로젝트 검색 범위 토글 (Scope Switch)", "S09", "개인 공간 컨트롤",
         "• 토글 스위치 (ON: 검정색 배경 #3A3733 / OFF: 회색 배경 #DDD9D3)\n• 검색 범위 상태 배지 (ON · 검색 범위 / OFF · 제외)",
         "• 스위치 클릭으로 해당 프로젝트 전용 매뉴얼 검색 스코프 활성화/비활성화\n• ON 설정 시 메인 대화창 상단에 '범위: [프로젝트명]' 칩 자동 연동",
         "FN-PWS-02 전용 매뉴얼 맞춤형 RAG 검색 범위 격리"),

        ("임계값 슬라이더 컨트롤 (Threshold Slider)", "S11", "관리자 설정 위젯",
         "• 질의 유형 레이블 (SOP/매뉴얼 질의, 트러블슈팅 이력)\n• 현재 설정 수치 (0.70, 0.65)\n• 슬라이더 트랙 및 드래그 핸들러",
         "• 핸들러를 좌우로 드래그하여 유사도 임계값 실시간 변경\n• 변경 즉시 Vector DB 검색 엔진에 반영되어 Hallucination 방지 기준 조정",
         "5.7 지식 노후화 및 신뢰도 관리 정책 연동"),

        ("계정 상태 및 잠금 해제 패널 (Account Action Panel)", "S13", "관리자 계정 관리",
         "• RBAC 권한 토글 (User / Admin)\n• 실패 횟수 카운터 (0 / 5회)\n• 잠금 알림 경고 카드 (5회 초과 계정 목록)\n• [잠금 해제] / [비밀번호 초기화] 버튼",
         "• [잠금 해제] 클릭 시 실패 카운트가 0으로 초기화되고 계정 상태가 '정상'으로 즉시 복구됨\n• [비밀번호 초기화] 클릭 시 임시 비밀번호 발급 및 변경 강제 플래그 설정",
         "SR-003, SR-004 계정 잠금 정책 및 권한 제어")
    ]
    
    for c_idx, h in enumerate(comp_headers, start=2):
        cell = ws_comp.cell(row=5, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
        
    for r_idx, row in enumerate(comp_data, start=6):
        fill_color = zebra_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        for c_idx, val in enumerate(row, start=2):
            cell = ws_comp.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.border = cell_border
            cell.fill = fill_color
            
            if c_idx in [2, 4]: # 컴포넌트명, 유형
                cell.font = cell_font_bold
                cell.alignment = align_left if c_idx == 2 else align_center
            elif c_idx == 3: # 적용 화면
                cell.font = Font(name="Consolas", size=9.5, color="1E3A8A")
                cell.alignment = align_center
            else: # 세부 구성, 인터랙션, 시스템 규칙
                cell.alignment = align_left
                
    ws_comp.column_dimensions['A'].width = 3
    ws_comp.column_dimensions['B'].width = 30  # Component Name
    ws_comp.column_dimensions['C'].width = 25  # Screens
    ws_comp.column_dimensions['D'].width = 22  # Type
    ws_comp.column_dimensions['E'].width = 48  # UI Elements
    ws_comp.column_dimensions['F'].width = 48  # Interaction
    ws_comp.column_dimensions['G'].width = 35  # Rules

    # Save to file
    output_filename = "Factory_Operation_AI_Agent_IA.xlsx"
    wb.save(output_filename)
    print(f"Successfully generated IA Excel file: {output_filename}")

if __name__ == "__main__":
    create_ia_excel()
